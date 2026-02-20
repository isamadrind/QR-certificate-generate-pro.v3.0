"""
╔═══════════════════════════════════════════════════════════════════════╗
║         QR Certificate System v4.0                                    ║
║         Developed by: Abdul Samad | SBBU Nawabshah                    ║
╠═══════════════════════════════════════════════════════════════════════╣
║  INSTALL:                                                             ║
║    pip install streamlit pillow qrcode[pil] reportlab openpyxl pandas ║
║  RUN:  streamlit run app.py                                           ║
╚═══════════════════════════════════════════════════════════════════════╝

FIXES in v4.0:
  ✅ Data CSV mein save hoti hai — student + admin same data dekhte hain
  ✅ URL ek baar set karo — config file mein save rehta hai, dobara nahi maangta
  ✅ Student form submit hone ka confirmation dono jagah dikh ta hai
"""

import streamlit as st
from PIL import Image, ImageDraw, ImageFont
import qrcode
import io, zipfile, csv, os, json
import pandas as pd
import openpyxl
from openpyxl.styles import Font as XFont, PatternFill, Alignment, Border, Side
from openpyxl.utils import get_column_letter
from reportlab.pdfgen import canvas as pdf_canvas
from reportlab.lib.pagesizes import landscape, A4
from reportlab.lib.utils import ImageReader
from datetime import datetime

# ─────────────────────────────────────────────────────────────────
#  FILE PATHS  (server par save hote hain)
# ─────────────────────────────────────────────────────────────────
DATA_FILE   = "registrations.csv"   # sab registrations yahan save hongi
CONFIG_FILE = "config.json"         # event info + URL yahan save hogi

CSV_HEADERS = ["name","roll_no","department","batch","category","event","date","time"]

# ─────────────────────────────────────────────────────────────────
#  CSV DATABASE FUNCTIONS
# ─────────────────────────────────────────────────────────────────
def save_registration(rec: dict):
    """Ek registration CSV mein append karo."""
    file_exists = os.path.exists(DATA_FILE)
    with open(DATA_FILE, "a", newline="", encoding="utf-8") as f:
        writer = csv.DictWriter(f, fieldnames=CSV_HEADERS)
        if not file_exists:
            writer.writeheader()
        writer.writerow({k: rec.get(k, "") for k in CSV_HEADERS})

def load_registrations() -> list:
    """Sab registrations CSV se load karo."""
    if not os.path.exists(DATA_FILE):
        return []
    try:
        with open(DATA_FILE, "r", encoding="utf-8") as f:
            reader = csv.DictReader(f)
            return [row for row in reader]
    except Exception:
        return []

def clear_registrations():
    """Sab data delete karo."""
    if os.path.exists(DATA_FILE):
        os.remove(DATA_FILE)

def save_config(cfg: dict):
    """Config JSON mein save karo."""
    with open(CONFIG_FILE, "w", encoding="utf-8") as f:
        json.dump(cfg, f, ensure_ascii=False, indent=2)

def load_config() -> dict:
    """Config load karo, agar nahi hai to defaults."""
    defaults = {
        "event_name":  "Certificate of Participation",
        "event_date":  datetime.now().strftime("%Y-%m-%d"),
        "event_venue": "",
        "event_topic": "",
        "organizer":   "",
        "categories":  "Participant,Management",
        "app_url":     "",
    }
    if not os.path.exists(CONFIG_FILE):
        return defaults
    try:
        with open(CONFIG_FILE, "r", encoding="utf-8") as f:
            saved = json.load(f)
        defaults.update(saved)
        return defaults
    except Exception:
        return defaults

# ─────────────────────────────────────────────────────────────────
#  PAGE CONFIG
# ─────────────────────────────────────────────────────────────────
st.set_page_config(
    page_title="QR Certificate Generator Pro",
    page_icon="🎓",
    layout="wide",
    initial_sidebar_state="expanded"
)

# ─────────────────────────────────────────────────────────────────
#  CSS
# ─────────────────────────────────────────────────────────────────
st.markdown("""
<style>
.stApp{background:linear-gradient(135deg,#0b132b 0%,#1c2541 100%);}
section[data-testid="stSidebar"]{background:#1e1b4b!important;}
section[data-testid="stSidebar"] *{color:#7ecefd!important;}
h1{color:#ffd159!important;text-align:center;}
h2,h3{color:#7ecefd!important;}
label,.stTextInput label,.stSelectbox label,
.stSlider label,.stTextArea label{color:#7ecefd!important;font-weight:600;}
p{color:#c5d8f0;}
.stTextInput>div>div>input,
.stNumberInput>div>div>input,
.stTextArea textarea{
    background:#0d1b35!important;color:white!important;
    border:1.5px solid #7ecefd55!important;border-radius:8px!important;
    font-size:1rem!important;}
.stTextInput>div>div>input:focus,.stTextArea textarea:focus{
    border-color:#7ecefd!important;box-shadow:0 0 0 2px #7ecefd33!important;}
.stSelectbox>div>div{
    background:#0d1b35!important;color:white!important;
    border:1.5px solid #7ecefd55!important;border-radius:8px!important;}
.stButton>button{
    background:linear-gradient(90deg,#2e6bef,#7ecefd)!important;
    color:white!important;border:none!important;border-radius:10px!important;
    font-weight:bold!important;font-size:1rem!important;padding:.6rem 1.2rem!important;
    transition:all .2s!important;}
.stButton>button:hover{opacity:.85!important;transform:scale(1.01)!important;}
.card{background:rgba(20,30,70,.92);border:1px solid #7ecefd33;
      border-radius:16px;padding:24px;margin:10px 0;}
.card-green{background:rgba(10,60,40,.9);border:1px solid #2ecc7166;
            border-radius:14px;padding:20px;margin:12px 0;}
.card-warn{background:rgba(80,40,0,.85);border:1px solid #f39c1266;
           border-radius:14px;padding:16px;margin:10px 0;}
.card-blue{background:rgba(10,40,80,.9);border:1px solid #3498db66;
           border-radius:14px;padding:18px;margin:10px 0;}
[data-testid="stMetricValue"]{color:#ffd159!important;font-size:2rem!important;}
[data-testid="stMetricLabel"]{color:#7ecefd!important;}
.stTabs [data-baseweb="tab"]{color:#7ecefd;background:#1e1b4b;border-radius:8px 8px 0 0;font-weight:600;}
.stTabs [aria-selected="true"]{background:#2e6bef!important;color:white!important;}
.stDataFrame{border-radius:10px;overflow:hidden;}
hr{border-color:#7ecefd22!important;}
</style>
""", unsafe_allow_html=True)

# ─────────────────────────────────────────────────────────────────
#  LOAD CONFIG (startup par ek baar)
# ─────────────────────────────────────────────────────────────────
cfg_file = load_config()

# ─────────────────────────────────────────────────────────────────
#  SESSION STATE
# ─────────────────────────────────────────────────────────────────
SESS_DEFAULTS = {
    "admin_auth":     False,
    "admin_password": "admin123",
    "template_bytes": None,
    "qr_data":        None,
    # from config file
    "event_name":     cfg_file["event_name"],
    "event_date":     cfg_file["event_date"],
    "event_venue":    cfg_file["event_venue"],
    "event_topic":    cfg_file["event_topic"],
    "organizer":      cfg_file["organizer"],
    "categories":     cfg_file["categories"],
    "app_url":        cfg_file["app_url"],
    # text settings
    "text_x":         50,
    "text_y":         60,
    "font_size":      72,
    "text_color":     "#1a1a1a",
    "selected_font":  "Arial Bold",
}
for k, v in SESS_DEFAULTS.items():
    if k not in st.session_state:
        st.session_state[k] = v

# ─────────────────────────────────────────────────────────────────
#  FONTS
# ─────────────────────────────────────────────────────────────────
FONTS = {
    "Arial Regular":          ["arial.ttf",        "DejaVuSans.ttf"],
    "Arial Bold":             ["arialbd.ttf",       "DejaVuSans-Bold.ttf"],
    "Arial Italic":           ["ariali.ttf",        "DejaVuSans-Oblique.ttf"],
    "Arial Bold Italic":      ["arialbi.ttf",       "DejaVuSans-BoldOblique.ttf"],
    "Calibri Regular":        ["calibri.ttf",       "DejaVuSans.ttf"],
    "Calibri Bold":           ["calibrib.ttf",      "DejaVuSans-Bold.ttf"],
    "Calibri Italic":         ["calibrii.ttf",      "DejaVuSans-Oblique.ttf"],
    "Tahoma Regular":         ["tahoma.ttf",        "DejaVuSans.ttf"],
    "Tahoma Bold":            ["tahomabd.ttf",      "DejaVuSans-Bold.ttf"],
    "Verdana Regular":        ["verdana.ttf",       "DejaVuSans.ttf"],
    "Verdana Bold":           ["verdanab.ttf",      "DejaVuSans-Bold.ttf"],
    "Trebuchet MS":           ["trebuc.ttf",        "DejaVuSans.ttf"],
    "Trebuchet Bold":         ["trebucbd.ttf",      "DejaVuSans-Bold.ttf"],
    "Segoe UI":               ["segoeui.ttf",       "DejaVuSans.ttf"],
    "Segoe UI Bold":          ["segoeuib.ttf",      "DejaVuSans-Bold.ttf"],
    "Segoe UI Light":         ["segoeuil.ttf",      "DejaVuSans.ttf"],
    "Times New Roman":        ["times.ttf",         "DejaVuSerif.ttf"],
    "Times New Roman Bold":   ["timesbd.ttf",       "DejaVuSerif-Bold.ttf"],
    "Times New Roman Italic": ["timesi.ttf",        "DejaVuSerif-Italic.ttf"],
    "Times NR Bold Italic":   ["timesbi.ttf",       "DejaVuSerif-BoldItalic.ttf"],
    "Georgia Regular":        ["georgia.ttf",       "DejaVuSerif.ttf"],
    "Georgia Bold":           ["georgiab.ttf",      "DejaVuSerif-Bold.ttf"],
    "Georgia Italic":         ["georgiai.ttf",      "DejaVuSerif-Italic.ttf"],
    "Palatino Linotype":      ["pala.ttf",          "DejaVuSerif.ttf"],
    "Palatino Bold":          ["palab.ttf",         "DejaVuSerif-Bold.ttf"],
    "Book Antiqua":           ["bkant.ttf",         "DejaVuSerif.ttf"],
    "Garamond":               ["GARA.TTF",          "DejaVuSerif.ttf"],
    "Garamond Bold":          ["GARABD.TTF",        "DejaVuSerif-Bold.ttf"],
    "Courier New":            ["cour.ttf",          "DejaVuSansMono.ttf"],
    "Courier New Bold":       ["courbd.ttf",        "DejaVuSansMono-Bold.ttf"],
    "Courier Italic":         ["couri.ttf",         "DejaVuSansMono-Oblique.ttf"],
    "Consolas":               ["consola.ttf",       "DejaVuSansMono.ttf"],
    "Consolas Bold":          ["consolab.ttf",      "DejaVuSansMono-Bold.ttf"],
    "Lucida Console":         ["lucon.ttf",         "DejaVuSansMono.ttf"],
    "Century Gothic":         ["GOTHIC.TTF",        "DejaVuSans.ttf"],
    "Century Gothic Bold":    ["GOTHICB.TTF",       "DejaVuSans-Bold.ttf"],
    "Century Gothic Italic":  ["GOTHICI.TTF",       "DejaVuSans-Oblique.ttf"],
    "Impact":                 ["impact.ttf",        "DejaVuSans-Bold.ttf"],
    "Franklin Gothic":        ["framd.ttf",         "DejaVuSans-Bold.ttf"],
    "Candara Regular":        ["Candara.ttf",       "DejaVuSans.ttf"],
    "Candara Bold":           ["Candarab.ttf",      "DejaVuSans-Bold.ttf"],
    "Corbel Regular":         ["corbel.ttf",        "DejaVuSans.ttf"],
    "Corbel Bold":            ["corbelb.ttf",       "DejaVuSans-Bold.ttf"],
    "Rockwell":               ["ROCK.TTF",          "DejaVuSerif.ttf"],
    "Rockwell Bold":          ["ROCKB.TTF",         "DejaVuSerif-Bold.ttf"],
    "Brush Script MT":        ["BRUSHSCI.TTF",      "DejaVuSerif-Italic.ttf"],
    "Lucida Handwriting":     ["lhandw.ttf",        "DejaVuSerif-Italic.ttf"],
    "Lucida Calligraphy":     ["LCALLIG.TTF",       "DejaVuSerif-Italic.ttf"],
    "Comic Sans MS":          ["comic.ttf",         "DejaVuSans.ttf"],
    "Comic Sans Bold":        ["comicbd.ttf",       "DejaVuSans-Bold.ttf"],
    "DejaVu Sans":            ["DejaVuSans.ttf",          "DejaVuSans.ttf"],
    "DejaVu Sans Bold":       ["DejaVuSans-Bold.ttf",     "DejaVuSans-Bold.ttf"],
    "DejaVu Serif":           ["DejaVuSerif.ttf",         "DejaVuSerif.ttf"],
    "DejaVu Serif Bold":      ["DejaVuSerif-Bold.ttf",    "DejaVuSerif-Bold.ttf"],
    "DejaVu Mono":            ["DejaVuSansMono.ttf",      "DejaVuSansMono.ttf"],
    "DejaVu Mono Bold":       ["DejaVuSansMono-Bold.ttf", "DejaVuSansMono-Bold.ttf"],
}

FONT_CATS = {
    "🔤 Sans Serif":    [k for k in FONTS if any(x in k for x in ["Arial","Calibri","Tahoma","Verdana","Trebuchet","Segoe"])],
    "📜 Serif/Formal":  [k for k in FONTS if any(x in k for x in ["Times","Georgia","Palatino","Book","Garamond"])],
    "💻 Monospace":     [k for k in FONTS if any(x in k for x in ["Courier","Consolas","Lucida Console"])],
    "✨ Display":       [k for k in FONTS if any(x in k for x in ["Century","Impact","Franklin","Candara","Corbel","Rockwell"])],
    "🖋️ Script":       [k for k in FONTS if any(x in k for x in ["Brush","Handwriting","Calligraphy","Comic"])],
    "🛡️ Fallback":     [k for k in FONTS if "DejaVu" in k],
}

# ─────────────────────────────────────────────────────────────────
#  CORE HELPERS
# ─────────────────────────────────────────────────────────────────
def load_font(name: str, size: int) -> ImageFont.ImageFont:
    for path in FONTS.get(name, ["DejaVuSans-Bold.ttf"]):
        try:
            return ImageFont.truetype(path, size)
        except Exception:
            continue
    return ImageFont.load_default()

def hex_to_rgba(h: str, alpha: int = 255):
    h = h.lstrip("#")
    return (int(h[0:2],16), int(h[2:4],16), int(h[4:6],16), alpha)

def generate_cert(name: str, template: bytes, cfg: dict) -> bytes:
    img   = Image.open(io.BytesIO(template)).convert("RGBA")
    w, h  = img.size
    font  = load_font(cfg["font"], cfg["size"])
    px    = int(w * cfg["x"] / 100)
    py    = int(h * cfg["y"] / 100)
    layer = Image.new("RGBA", img.size, (255,255,255,0))
    draw  = ImageDraw.Draw(layer)
    bbox  = draw.textbbox((0,0), name, font=font)
    tw, th = bbox[2]-bbox[0], bbox[3]-bbox[1]
    draw.text((px-tw//2, py-th//2), name,
              font=font, fill=hex_to_rgba(cfg["color"]))
    out = Image.alpha_composite(img, layer).convert("RGB")
    buf = io.BytesIO()
    out.save(buf, format="PNG", dpi=(300,300))
    return buf.getvalue()

def cert_to_pdf(png: bytes, name: str) -> bytes:
    buf    = io.BytesIO()
    pw, ph = landscape(A4)
    c      = pdf_canvas.Canvas(buf, pagesize=(pw,ph))
    img    = Image.open(io.BytesIO(png)).convert("RGB")
    iw, ih = img.size
    sc     = min(pw/iw, ph/ih)
    nw, nh = iw*sc, ih*sc
    tmp    = io.BytesIO()
    img.save(tmp, format="PNG"); tmp.seek(0)
    c.drawImage(ImageReader(tmp),(pw-nw)/2,(ph-nh)/2,nw,nh,mask="auto")
    c.setFont("Helvetica-Bold",9)
    c.setFillColorRGB(.5,.5,.5)
    c.drawCentredString(pw/2,14,
        f"{name}  |  {st.session_state.event_name}  |  {datetime.now().strftime('%Y-%m-%d')}")
    c.save()
    return buf.getvalue()

def make_qr(url: str) -> bytes:
    qr = qrcode.QRCode(version=1,
                       error_correction=qrcode.constants.ERROR_CORRECT_H,
                       box_size=10, border=4)
    qr.add_data(url)
    qr.make(fit=True)
    buf = io.BytesIO()
    qr.make_image(fill_color="#0b132b", back_color="white").save(buf, format="PNG")
    return buf.getvalue()

def cur_cfg() -> dict:
    return {
        "x":    st.session_state.text_x,
        "y":    st.session_state.text_y,
        "size": st.session_state.font_size,
        "color":st.session_state.text_color,
        "font": st.session_state.selected_font,
    }

def build_excel(regs: list) -> bytes:
    wb  = openpyxl.Workbook()
    hf  = PatternFill("solid", fgColor="1E1B4B")
    hf2 = PatternFill("solid", fgColor="0B132B")
    hfn = XFont(bold=True, color="FFFFFF", size=12)
    bdr = Border(bottom=Side(style="thin", color="334466"))

    ws = wb.active
    ws.title = "Registrations"
    ws.merge_cells("A1:H1")
    t = ws["A1"]
    t.value = f"  {st.session_state.event_name} — Registration Data"
    t.font  = XFont(bold=True, color="FFD159", size=14)
    t.fill  = hf2
    t.alignment = Alignment(horizontal="center", vertical="center")
    ws.row_dimensions[1].height = 34

    ws.merge_cells("A2:H2")
    info = ws["A2"]
    try:
        d   = datetime.strptime(st.session_state.event_date, "%Y-%m-%d")
        day = d.strftime("%A")
    except Exception:
        day = ""
    info.value = (f"Date: {st.session_state.event_date} ({day})  |  "
                  f"Venue: {st.session_state.event_venue}  |  "
                  f"Topic: {st.session_state.event_topic}  |  "
                  f"Organizer: {st.session_state.organizer}  |  "
                  f"Total: {len(regs)}")
    info.font  = XFont(color="7ECEFD", size=10)
    info.fill  = hf
    info.alignment = Alignment(horizontal="center")
    ws.row_dimensions[2].height = 18

    cols_info = [("#",5),("Full Name",28),("Roll No",16),
                 ("Department",24),("Batch",14),("Category",16),
                 ("Date",14),("Time",10)]
    for ci,(h,w) in enumerate(cols_info,1):
        cell = ws.cell(row=3,column=ci,value=h)
        cell.font = hfn; cell.fill = hf
        cell.alignment = Alignment(horizontal="center")
        ws.column_dimensions[get_column_letter(ci)].width = w
    ws.row_dimensions[3].height = 22

    for ri,rec in enumerate(regs,4):
        alt = PatternFill("solid", fgColor="0F1B35" if ri%2==0 else "1A2550")
        vals = [ri-3,rec.get("name",""),rec.get("roll_no",""),
                rec.get("department",""),rec.get("batch",""),
                rec.get("category",""),rec.get("date",""),rec.get("time","")]
        for ci,val in enumerate(vals,1):
            c = ws.cell(row=ri,column=ci,value=val)
            c.font = XFont(color="E0E0E0",size=11)
            c.fill = alt; c.border = bdr
            c.alignment = Alignment(
                horizontal="center" if ci in [1,6,7,8] else "left",
                vertical="center")
        ws.row_dimensions[ri].height = 20

    ws2 = wb.create_sheet("Category Summary")
    ws2.merge_cells("A1:C1")
    t2 = ws2["A1"]; t2.value = "Category-wise Summary"
    t2.font = XFont(bold=True,color="FFD159",size=13)
    t2.fill = hf2; t2.alignment = Alignment(horizontal="center")
    ws2.row_dimensions[1].height = 28
    for ci,h in enumerate(["Category","Count","Members (Roll No)"],1):
        c = ws2.cell(row=2,column=ci,value=h)
        c.font = hfn; c.fill = hf
        c.alignment = Alignment(horizontal="center")
    cats: dict = {}
    for rec in regs:
        cats.setdefault(rec.get("category","Other"),[]).append(
            f"{rec.get('name','')} [{rec.get('roll_no','')}]")
    for ri,(cat,names) in enumerate(cats.items(),3):
        ws2.cell(row=ri,column=1,value=cat).font = XFont(bold=True,color="FFD159")
        ws2.cell(row=ri,column=2,value=len(names)).font = XFont(color="E0E0E0")
        ws2.cell(row=ri,column=3,value=", ".join(names)).font = XFont(color="E0E0E0")
        for col in range(1,4):
            ws2.cell(row=ri,column=col).fill = hf
    ws2.column_dimensions["A"].width = 20
    ws2.column_dimensions["B"].width = 10
    ws2.column_dimensions["C"].width = 80

    buf = io.BytesIO()
    wb.save(buf)
    return buf.getvalue()

def save_all_settings():
    """Sidebar settings ko config file mein save karo."""
    save_config({
        "event_name":  st.session_state.event_name,
        "event_date":  st.session_state.event_date,
        "event_venue": st.session_state.event_venue,
        "event_topic": st.session_state.event_topic,
        "organizer":   st.session_state.organizer,
        "categories":  st.session_state.categories,
        "app_url":     st.session_state.app_url,
    })

# ═════════════════════════════════════════════════════════════════
#  ROUTING
# ═════════════════════════════════════════════════════════════════
qp   = st.query_params
page = qp.get("page", "admin")

# ═════════════════════════════════════════════════════════════════
#  STUDENT FORM PAGE  — config file se data parhta hai
# ═════════════════════════════════════════════════════════════════
if page == "form":
    # Config file se settings load karo (admin ne jo set kiya)
    cfg = load_config()
    event    = cfg.get("event_name", "Certificate Event")
    cats_str = cfg.get("categories", "Participant,Management")
    cats     = [c.strip() for c in cats_str.split(",") if c.strip()]

    st.markdown(f"""
    <div style="text-align:center;padding:30px 0 10px;">
      <div style="font-size:3.5rem;">🎓</div>
      <h1 style="color:#ffd159;font-size:2rem;margin:8px 0;">{event}</h1>
      <p style="color:#7ecefd;font-size:1rem;margin:4px 0;">
        {cfg.get('event_venue','')}
        {"  |  " + cfg.get('event_date','') if cfg.get('event_date') else ''}
      </p>
      <p style="color:#7ecefd88;font-size:.9rem;">
        Registration Form — Apni details bharein
      </p>
    </div>
    """, unsafe_allow_html=True)
    st.markdown("---")

    # Check if already submitted (same session)
    if st.session_state.get("form_submitted"):
        rec = st.session_state.get("last_submission", {})
        st.markdown(f"""
        <div class="card-green">
          <h3 style="color:#2ecc71;text-align:center;margin:0 0 14px;">
            ✅ Aapki Registration Ho Chuki Hai!
          </h3>
          <table style="width:100%;color:#c5d8f0;font-size:1.05rem;border-collapse:collapse;">
            <tr><td style="padding:6px 0;width:40%;">👤 <b>Name</b></td>
                <td style="color:white;font-weight:bold;">{rec.get('name','')}</td></tr>
            <tr><td style="padding:6px 0;">🔢 <b>Roll No</b></td>
                <td style="color:white;">{rec.get('roll_no','')}</td></tr>
            <tr><td style="padding:6px 0;">🏫 <b>Department</b></td>
                <td style="color:white;">{rec.get('department','')}</td></tr>
            <tr><td style="padding:6px 0;">📅 <b>Batch</b></td>
                <td style="color:white;">{rec.get('batch','')}</td></tr>
            <tr><td style="padding:6px 0;">🏷️ <b>Category</b></td>
                <td style="color:white;">{rec.get('category','')}</td></tr>
            <tr><td style="padding:6px 0;">🕐 <b>Time</b></td>
                <td style="color:white;">{rec.get('time','')}</td></tr>
          </table>
          <p style="color:#7ecefd;text-align:center;margin:14px 0 0;font-size:1rem;">
            ✅ Aapka data save ho gaya hai. Shukriya!
          </p>
        </div>
        """, unsafe_allow_html=True)

        if st.button("🔄 New Registration", use_container_width=True):
            st.session_state.form_submitted = False
            st.session_state.last_submission = {}
            st.rerun()

    else:
        st.markdown('<div class="card">', unsafe_allow_html=True)

        c1, c2 = st.columns(2)
        with c1:
            name   = st.text_input("👤 Poora Naam / Full Name ✱",
                                   placeholder="Abdul Samad", key="f_name")
            dept   = st.text_input("🏫 Department ✱",
                                   placeholder="Computer Science", key="f_dept")
        with c2:
            rollno = st.text_input("🔢 Roll No ✱",
                                   placeholder="24-BSCS-18", key="f_roll")
            batch  = st.text_input("📅 Batch / Year ✱",
                                   placeholder="2024", key="f_batch")

        category = st.selectbox("🏷️ Category — Choose accordingly ✱", cats, key="f_cat")

        st.markdown("---")
        submitted = st.button("✅  Submit Now", use_container_width=True)

        if submitted:
            n = name.strip(); r = rollno.strip()
            d = dept.strip(); b = batch.strip()
            missing = [f for f,v in
                      [("Full Name",n),("Roll No",r),("Department",d),("Batch",b)]
                      if not v]
            if missing:
                st.error("❌ Yeh fields zaroori hain: **" + "  |  ".join(missing) + "**")
            else:
                now = datetime.now()
                rec = {
                    "name": n, "roll_no": r, "department": d,
                    "batch": b, "category": category,
                    "event": event,
                    "date":  now.strftime("%Y-%m-%d"),
                    "time":  now.strftime("%H:%M:%S"),
                }
                # ✅ CSV mein save karo — permanent storage
                save_registration(rec)
                # Session mein bhi rakho for confirmation screen
                st.session_state.form_submitted  = True
                st.session_state.last_submission = rec
                st.rerun()

        st.markdown('</div>', unsafe_allow_html=True)

    st.markdown(
        '<p style="text-align:center;color:#7ecefd33;font-size:.8rem;margin-top:24px;">'
        'Developed by Abdul Samad — SBBU Nawabshah</p>',
        unsafe_allow_html=True)
    st.stop()

# ═════════════════════════════════════════════════════════════════
#  ADMIN PAGE
# ═════════════════════════════════════════════════════════════════
st.markdown("# 🎓 QR Certificate Generator Pro V3.0")
st.markdown(
    '<p style="text-align:center;color:#7ecefd;">'
    'Abdul Samad | Shaheed Benazir Bhutto University Nawabshah</p>',
    unsafe_allow_html=True)
st.markdown("---")

# ── Auth ─────────────────────────────────────────────────────────
if not st.session_state.admin_auth:
    _, col, _ = st.columns([1, 2, 1])
    with col:
        st.markdown('<div class="card">', unsafe_allow_html=True)
        st.markdown("### 🔐 Admin Login")
        pwd = st.text_input("Password", type="password")
        if st.button("🔓 Login", use_container_width=True):
            if pwd == st.session_state.admin_password:
                st.session_state.admin_auth = True
                st.rerun()
            else:
                st.error("❌ Wrong password!")
        st.caption("Password is required")
        st.markdown('</div>', unsafe_allow_html=True)
    st.stop()

# ── Sidebar ───────────────────────────────────────────────────────
with st.sidebar:
    st.markdown("## 📋 Event Settings")
    st.session_state.event_name  = st.text_input("Event Name",         st.session_state.event_name)
    st.session_state.event_topic = st.text_input("Topic",              st.session_state.event_topic)
    st.session_state.event_date  = st.text_input("Date (YYYY-MM-DD)",  st.session_state.event_date)
    st.session_state.event_venue = st.text_input("Venue",              st.session_state.event_venue)
    st.session_state.organizer   = st.text_input("Organizer",          st.session_state.organizer)
    st.session_state.categories  = st.text_input("Categories (comma)", st.session_state.categories)

    st.markdown("---")
    st.markdown("## 🌐 App URL")
    st.session_state.app_url = st.text_input(
        "Deployed App URL",
        value=st.session_state.app_url,
        placeholder="https://yourname-app.streamlit.app",
        help="Ek baar set karo — save ho jata hai, dobara nahi maangta")

    if st.button("💾 Save Settings", use_container_width=True):
        save_all_settings()
        st.success("✅ Saved! QR ab naya URL use karega.")

    st.markdown("---")
    st.markdown("## 🎨 Certificate Text")
    st.session_state.font_size  = st.slider("Font Size",         20, 250, st.session_state.font_size)
    st.session_state.text_x    = st.slider("Horizontal % (←→)", 0,  100, st.session_state.text_x)
    st.session_state.text_y    = st.slider("Vertical %   (↑↓)", 0,  100, st.session_state.text_y)
    st.session_state.text_color= st.color_picker("Text Color",   st.session_state.text_color)

    st.markdown("---")
    st.markdown("## 🔤 Font Select")
    search_q = st.text_input("🔍 Font Search...", placeholder="e.g. bold, times, gothic")
    all_fonts = list(FONTS.keys())

    if search_q.strip():
        matched = [f for f in all_fonts if search_q.strip().lower() in f.lower()]
        if matched:
            st.caption(f"{len(matched)} fonts mile")
            idx = matched.index(st.session_state.selected_font) \
                  if st.session_state.selected_font in matched else 0
            st.session_state.selected_font = st.selectbox(
                "Results:", matched, index=idx, key="fs_search")
        else:
            st.warning("Koi font nahi mila")
    else:
        for cat_lbl, cat_fonts in FONT_CATS.items():
            if not cat_fonts: continue
            expand = "Sans" in cat_lbl
            with st.expander(cat_lbl, expanded=expand):
                for fn in cat_fonts:
                    lbl = ("✅ " if st.session_state.selected_font == fn else "") + fn
                    if st.button(lbl, key=f"fb_{fn}", use_container_width=True):
                        st.session_state.selected_font = fn
                        st.rerun()

    st.markdown(f"**Selected:** `{st.session_state.selected_font}`")
    st.markdown("---")
    with st.expander("🔑 Change Password"):
        np_ = st.text_input("New Password", type="password")
        if st.button("Update") and np_:
            st.session_state.admin_password = np_
            st.success("✅ Updated!")
    if st.button("🚪 Logout"):
        st.session_state.admin_auth = False
        st.rerun()

# ── TABS ─────────────────────────────────────────────────────────
tab1, tab2, tab3, tab4, tab5 = st.tabs([
    "🔳 QR Generate",
    "📊 Registered Data",
    "🖼️ Template & Preview",
    "🚀 Generate Certificates",
    "☁️ GitHub Guide",
])

# ═══════════════════════════════════════
#  TAB 1 — QR Generate
# ═══════════════════════════════════════
with tab1:
    cl, cr = st.columns(2)

    with cl:
        st.markdown('<div class="card">', unsafe_allow_html=True)
        st.markdown("### 🔳 QR Code")

        # URL already saved — auto load
        saved_url = st.session_state.app_url
        if saved_url:
            st.markdown(f"""
            <div class="card-blue">
              ✅ <b>Saved URL:</b><br>
              <code style="color:#ffd159;">{saved_url}</code><br><br>
              URL sidebar mein save hai — bar bar likhne ki zaroorat nahi!
            </div>
            """, unsafe_allow_html=True)

            ev   = st.session_state.event_name.replace(" ","%20")
            cats = st.session_state.categories.replace(" ","%20")
            qr_url = f"{saved_url.rstrip('/')}/?page=form"

            if st.button("🔳 QR Generate / Refresh Karein", use_container_width=True):
                st.session_state.qr_data = make_qr(qr_url)
                st.success("✅ QR ready!")

            # Auto-generate on load if not already
            if not st.session_state.qr_data:
                st.session_state.qr_data = make_qr(qr_url)

            if st.session_state.qr_data:
                st.image(st.session_state.qr_data, width=250,
                         caption="Yeh QR print karo → event mein lagao")
                st.download_button("⬇️ QR PNG Download",
                    data=st.session_state.qr_data,
                    file_name="registration_qr.png", mime="image/png",
                    use_container_width=True)
                st.code(qr_url, language=None)
        else:
            st.markdown("""
            <div class="card-warn">
              ⚠️ <b>URL abhi set nahi hai!</b><br><br>
              Sidebar mein <b>"App URL"</b> field mein apni Streamlit URL paste karo,
              phir <b>"Save Settings"</b> click karo.<br><br>
              Sirf ek baar karna hai — hamesha save rehega!
            </div>
            """, unsafe_allow_html=True)

        st.markdown('</div>', unsafe_allow_html=True)

    with cr:
        st.markdown('<div class="card">', unsafe_allow_html=True)
        st.markdown("### 📱 Student Experience")
        st.markdown("""
Student QR scan karne ke baad **yeh form dekhega:**

| Field | Example |
|-------|---------|
| 👤 Full Name | Abdul Samad |
| 🔢 Roll No | 24-BSCS-18 |
| 🏫 Department | Computer Science |
| 📅 Batch | 2024 |
| 🏷️ Category | Participant |

**Submit** hone ke baad:
- ✅ Green confirmation screen dikh ta hai
- 📊 Data turant **Tab 2** mein aa jata hai
- 💾 CSV file mein permanent save ho jata hai
        """)
        st.markdown('</div>', unsafe_allow_html=True)

        # Manual add
        st.markdown('<div class="card">', unsafe_allow_html=True)
        st.markdown("### ✏️ Manual Entry")
        with st.form("manual_reg_form"):
            mc1, mc2 = st.columns(2)
            with mc1:
                mn = st.text_input("👤 Full Name")
                md = st.text_input("🏫 Department")
            with mc2:
                mr = st.text_input("🔢 Roll No")
                mb = st.text_input("📅 Batch")
            cat_list = [c.strip() for c in
                        st.session_state.categories.split(",") if c.strip()]
            mc = st.selectbox("🏷️ Category", cat_list)
            if st.form_submit_button("➕ Add", use_container_width=True):
                if mn.strip() and mr.strip():
                    now = datetime.now()
                    rec = {
                        "name":mn.strip(),"roll_no":mr.strip(),
                        "department":md.strip(),"batch":mb.strip(),
                        "category":mc,"event":st.session_state.event_name,
                        "date":now.strftime("%Y-%m-%d"),
                        "time":now.strftime("%H:%M:%S"),
                    }
                    save_registration(rec)
                    st.success(f"✅ {mn.strip()} add ho gaya!")
                    st.rerun()
                else:
                    st.error("Naam aur Roll No zaroori hain!")
        st.markdown('</div>', unsafe_allow_html=True)

# ═══════════════════════════════════════
#  TAB 2 — Registered Data
# ═══════════════════════════════════════
with tab2:
    # Har baar CSV se fresh data load karo
    regs = load_registrations()
    st.markdown("### 📊 Registered Data")

    # Auto-refresh button
    if st.button("🔄 Refresh Data"):
        st.rerun()

    cat_list = [c.strip() for c in st.session_state.categories.split(",") if c.strip()]
    m_cols   = st.columns(len(cat_list) + 1)
    m_cols[0].metric("Total", len(regs))
    for i, cat in enumerate(cat_list):
        m_cols[i+1].metric(cat, sum(1 for r in regs if r.get("category","") == cat))

    st.markdown("---")

    if regs:
        df = pd.DataFrame(regs)
        rename = {"name":"Full Name","roll_no":"Roll No","department":"Department",
                  "batch":"Batch","category":"Category","event":"Event",
                  "date":"Date","time":"Time"}
        df = df.rename(columns={k:v for k,v in rename.items() if k in df.columns})

        filter_cat = st.selectbox("Category filter:",
                                  ["All"] + cat_list, key="flt_cat")
        df_show = df[df["Category"] == filter_cat] if filter_cat != "All" else df
        st.dataframe(df_show, use_container_width=True, height=380)

        st.markdown("---")
        c1, c2, c3 = st.columns(3)
        with c1:
            excel = build_excel(regs)
            st.download_button("📊 Excel Download",
                data=excel,
                file_name=f"{st.session_state.event_name.replace(' ','_')}_Data.xlsx",
                mime="application/vnd.openxmlformats-officedocument.spreadsheetml.sheet",
                use_container_width=True)
        with c2:
            txt = "\n".join(
                f"{r['name']} | {r['roll_no']} | {r['department']} | {r['batch']} | {r['category']}"
                for r in regs)
            st.download_button("📄 TXT Download",
                data=txt.encode(), file_name="registrations.txt",
                mime="text/plain", use_container_width=True)
        with c3:
            if st.button("🗑️ Sab Clear Karo", use_container_width=True):
                clear_registrations()
                st.success("✅ Data clear ho gaya!")
                st.rerun()
    else:
        st.info("📭 Koi registration nahi abhi. QR scan hone par yahan data aayega.")
        st.markdown("""
        <div class="card-blue">
        💡 <b>Tip:</b> Jab bhi student QR scan karke form submit kare,
        yahan <b>🔄 Refresh Data</b> button click karo — data dikh jayega!
        </div>
        """, unsafe_allow_html=True)

# ═══════════════════════════════════════
#  TAB 3 — Template & Preview
# ═══════════════════════════════════════
with tab3:
    cl, cr = st.columns([1, 1])

    with cl:
        st.markdown('<div class="card">', unsafe_allow_html=True)
        st.markdown("### 🖼️ Certificate Template")
        tpl = st.file_uploader("Template (.png/.jpg/.jpeg)",
                               type=["png","jpg","jpeg"])
        if tpl:
            st.session_state.template_bytes = tpl.read()
            img_tmp = Image.open(io.BytesIO(st.session_state.template_bytes))
            st.success(f"✅ {tpl.name}  —  {img_tmp.width}×{img_tmp.height}px")
        if st.session_state.template_bytes:
            st.image(st.session_state.template_bytes,
                     caption="Current Template", use_container_width=True)
        else:
            st.info("Template upload karo")
        st.markdown('</div>', unsafe_allow_html=True)

    with cr:
        st.markdown('<div class="card">', unsafe_allow_html=True)
        st.markdown("### 👁️ Live Preview")
        if st.session_state.template_bytes:
            st.markdown(
                f"**Font:** `{st.session_state.selected_font}` | "
                f"**Size:** `{st.session_state.font_size}` | "
                f"**Pos:** ({st.session_state.text_x}%, {st.session_state.text_y}%) | "
                f"**Color:** `{st.session_state.text_color}`")
            pname    = st.text_input("Preview naam:", value="Muhammad Ali Khan")
            png_prev = generate_cert(pname, st.session_state.template_bytes, cur_cfg())
            st.image(png_prev, use_container_width=True)
            pa, pb = st.columns(2)
            with pa:
                st.download_button("⬇️ PNG", png_prev,
                    file_name=f"Preview_{pname}.png", mime="image/png",
                    use_container_width=True)
            with pb:
                st.download_button("⬇️ PDF", cert_to_pdf(png_prev, pname),
                    file_name=f"Preview_{pname}.pdf", mime="application/pdf",
                    use_container_width=True)
        else:
            st.warning("⚠️ Pehle template upload karo")
        st.markdown('</div>', unsafe_allow_html=True)

    # Preview all
    regs = load_registrations()
    if st.session_state.template_bytes and regs:
        st.markdown("---")
        st.markdown("### 👁️ Sabke Certificates Preview")
        names_all = [r["name"] for r in regs]
        show_n = st.slider("Kitne preview?",
                           1, min(len(names_all), 30), min(6, len(names_all)))
        per_row = 3
        for i in range(0, show_n, per_row):
            row_n = names_all[i:i+per_row]
            cs    = st.columns(per_row)
            for ci, nm in enumerate(row_n):
                with cs[ci]:
                    pv = generate_cert(nm, st.session_state.template_bytes, cur_cfg())
                    st.image(pv, caption=nm, use_container_width=True)
                    st.download_button(f"⬇️ {nm[:16]}",
                        data=pv, file_name=f"{nm}.png",
                        mime="image/png", key=f"pv_{nm}_{i}_{ci}")

# ═══════════════════════════════════════
#  TAB 4 — Generate Certificates
# ═══════════════════════════════════════
with tab4:
    st.markdown("### 🚀 Certificates Generate Karein")
    regs = load_registrations()

    if not st.session_state.template_bytes:
        st.markdown('<div class="card-warn">⚠️ Pehle Tab 3 mein template upload karo!</div>',
                    unsafe_allow_html=True)
    elif not regs:
        st.markdown('<div class="card-warn">⚠️ Koi registration nahi hai. QR scan karwao pehle.</div>',
                    unsafe_allow_html=True)
    else:
        c1, c2, c3, c4 = st.columns(4)
        c1.metric("Total", len(regs))
        c2.metric("Font",  (st.session_state.selected_font[:14]+"..."
                            if len(st.session_state.selected_font)>14
                            else st.session_state.selected_font))
        c3.metric("Size",  st.session_state.font_size)
        c4.metric("Pos",   f"{st.session_state.text_x}%,{st.session_state.text_y}%")

        st.markdown("---")
        fc1, fc2 = st.columns(2)
        with fc1: do_png = st.checkbox("✅ PNG", value=True)
        with fc2: do_pdf = st.checkbox("✅ PDF", value=True)

        if st.button(f"🚀 Generate All {len(regs)} Certificates",
                     use_container_width=True):
            cfg_now = cur_cfg()
            prog    = st.progress(0)
            status  = st.empty()
            buf_zip = io.BytesIO()

            with zipfile.ZipFile(buf_zip, "w", zipfile.ZIP_DEFLATED) as zf:
                for i, rec in enumerate(regs):
                    nm  = rec["name"]
                    cat = rec.get("category","Other")
                    status.markdown(f"⏳ **{nm}** [{cat}]  ({i+1}/{len(regs)})")
                    png = generate_cert(nm, st.session_state.template_bytes, cfg_now)
                    if do_png: zf.writestr(f"PNG/{cat}/{nm}.png", png)
                    if do_pdf: zf.writestr(f"PDF/{cat}/{nm}.pdf", cert_to_pdf(png, nm))
                    prog.progress((i+1)/len(regs))

            status.success(f"✅ {len(regs)} certificates ready!")
            st.balloons()
            zname = f"{st.session_state.event_name.replace(' ','_')}_Certificates.zip"
            st.download_button(
                f"⬇️ Download All ({len(regs)}) — ZIP",
                data=buf_zip.getvalue(),
                file_name=zname, mime="application/zip",
                use_container_width=True)

# ═══════════════════════════════════════
#  TAB 5 — GitHub Guide
# ═══════════════════════════════════════
with tab5:
    st.markdown("""
<div class="card">

## ☁️ GitHub + Streamlit Cloud — Free Hosting

### Sirf 2 files upload karo GitHub par:
```
app.py
requirements.txt
```

### PowerShell commands:
```bash
cd d:/Avalon.AI
git add app.py requirements.txt
git commit -m "QR Certificate Generator Pro V3.0"
git push
```

### URL set karne ka tarika (sirf ek baar):
1. Deploy hone ke baad URL copy karo
2. Admin panel → Sidebar → **"App URL"** mein paste karo
3. **"Settings Save Karein"** click karo
4. QR Generate karo → Print karo ✅

**Ab dobara URL likhne ki zaroorat nahi — config.json mein save ho jata hai!**

### Update karne ka tarika:
```bash
git add app.py
git commit -m "update"
git push
```
Streamlit auto-update ho jata hai!

</div>
    """, unsafe_allow_html=True)

st.markdown("---")
st.markdown(
    '<p style="text-align:center;color:#7ecefd44;font-size:.85rem;">'
    '© QR Certificate System v3.0 | Abdul Samad | SBBU Nawabshah</p>',
    unsafe_allow_html=True)
